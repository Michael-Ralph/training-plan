import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def create_training_plan():
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Training Plan"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Training plan data
    training_weeks = [
        {
            "week": 1,
            "focus": "Foundation Building",
            "sessions": [
                {"day": "Monday", "type": "Strength", "duration": "45 min", "intensity": "Low", "notes": "Full body compound movements"},
                {"day": "Tuesday", "type": "Cardio", "duration": "30 min", "intensity": "Moderate", "notes": "Steady state running"},
                {"day": "Wednesday", "type": "Rest", "duration": "-", "intensity": "-", "notes": "Active recovery or complete rest"},
                {"day": "Thursday", "type": "Strength", "duration": "45 min", "intensity": "Low", "notes": "Upper body focus"},
                {"day": "Friday", "type": "Cardio", "duration": "25 min", "intensity": "Low", "notes": "Easy pace walk/jog"},
                {"day": "Saturday", "type": "Flexibility", "duration": "30 min", "intensity": "Low", "notes": "Yoga or stretching"},
                {"day": "Sunday", "type": "Rest", "duration": "-", "intensity": "-", "notes": "Complete rest"}
            ]
        },
        {
            "week": 2,
            "focus": "Strength Development",
            "sessions": [
                {"day": "Monday", "type": "Strength", "duration": "50 min", "intensity": "Moderate", "notes": "Progressive overload"},
                {"day": "Tuesday", "type": "Cardio", "duration": "35 min", "intensity": "Moderate", "notes": "Interval training"},
                {"day": "Wednesday", "type": "Rest", "duration": "-", "intensity": "-", "notes": "Light stretching"},
                {"day": "Thursday", "type": "Strength", "duration": "50 min", "intensity": "Moderate", "notes": "Lower body focus"},
                {"day": "Friday", "type": "Cardio", "duration": "30 min", "intensity": "Moderate", "notes": "Cross-training"},
                {"day": "Saturday", "type": "Strength", "duration": "40 min", "intensity": "Low", "notes": "Core and stability"},
                {"day": "Sunday", "type": "Rest", "duration": "-", "intensity": "-", "notes": "Complete rest"}
            ]
        },
        {
            "week": 3,
            "focus": "Intensity Building",
            "sessions": [
                {"day": "Monday", "type": "Strength", "duration": "55 min", "intensity": "High", "notes": "Heavy compound lifts"},
                {"day": "Tuesday", "type": "Cardio", "duration": "40 min", "intensity": "High", "notes": "HIIT training"},
                {"day": "Wednesday", "type": "Rest", "duration": "-", "intensity": "-", "notes": "Active recovery"},
                {"day": "Thursday", "type": "Strength", "duration": "55 min", "intensity": "High", "notes": "Push/pull focus"},
                {"day": "Friday", "type": "Cardio", "duration": "35 min", "intensity": "Moderate", "notes": "Steady state cardio"},
                {"day": "Saturday", "type": "Strength", "duration": "45 min", "intensity": "Moderate", "notes": "Functional movements"},
                {"day": "Sunday", "type": "Rest", "duration": "-", "intensity": "-", "notes": "Complete rest"}
            ]
        },
        {
            "week": 4,
            "focus": "Recovery & Assessment",
            "sessions": [
                {"day": "Monday", "type": "Strength", "duration": "40 min", "intensity": "Low", "notes": "Deload week - light weights"},
                {"day": "Tuesday", "type": "Cardio", "duration": "25 min", "intensity": "Low", "notes": "Easy recovery run"},
                {"day": "Wednesday", "type": "Rest", "duration": "-", "intensity": "-", "notes": "Complete rest"},
                {"day": "Thursday", "type": "Strength", "duration": "40 min", "intensity": "Low", "notes": "Movement quality focus"},
                {"day": "Friday", "type": "Cardio", "duration": "30 min", "intensity": "Low", "notes": "Light activity"},
                {"day": "Saturday", "type": "Assessment", "duration": "60 min", "intensity": "Moderate", "notes": "Fitness testing & measurements"},
                {"day": "Sunday", "type": "Rest", "duration": "-", "intensity": "-", "notes": "Complete rest"}
            ]
        }
    ]
    
    # Create title
    ws.merge_cells('A1:G1')
    ws['A1'] = "4-WEEK TRAINING PLAN"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws['A1'].alignment = center_alignment
    
    # Add program info
    current_row = 3
    info_data = [
        ["Program Duration:", "4 Weeks"],
        ["Start Date:", datetime.now().strftime("%B %d, %Y")],
        ["Goal:", "General Fitness & Strength Building"],
        ["Level:", "Beginner to Intermediate"]
    ]
    
    for info in info_data:
        ws[f'A{current_row}'] = info[0]
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = info[1]
        current_row += 1
    
    current_row += 1
    
    # Create training schedule
    for week_data in training_weeks:
        # Week header
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws[f'A{current_row}'] = f"WEEK {week_data['week']}: {week_data['focus'].upper()}"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = header_fill
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Column headers
        headers = ["Day", "Training Type", "Duration", "Intensity", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1
        
        # Training sessions
        for session in week_data['sessions']:
            row_data = [
                session['day'],
                session['type'],
                session['duration'],
                session['intensity'],
                session['notes']
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                if col == 1:  # Day column
                    cell.font = Font(bold=True)
                if col in [2, 3, 4]:  # Type, Duration, Intensity
                    cell.alignment = center_alignment
            current_row += 1
        
        current_row += 1  # Add space between weeks
    
    # Add progress tracking section
    ws.merge_cells(f'A{current_row}:G{current_row}')
    ws[f'A{current_row}'] = "PROGRESS TRACKING"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].alignment = center_alignment
    current_row += 2
    
    # Progress tracking headers
    progress_headers = ["Week", "Weight (lbs)", "Body Fat %", "Cardio Time", "Strength Level", "Notes"]
    for col, header in enumerate(progress_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_alignment
        cell.border = border
    current_row += 1
    
    # Progress tracking rows
    for week in range(1, 5):
        row_data = [f"Week {week}", "", "", "", "", ""]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border
            if col == 1:
                cell.font = Font(bold=True)
            else:
                cell.alignment = center_alignment
        current_row += 1
    
    # Adjust column widths
    column_widths = {
        'A': 12,  # Day
        'B': 15,  # Training Type
        'C': 12,  # Duration
        'D': 12,  # Intensity
        'E': 40,  # Notes
        'F': 12,  # Additional columns for progress
        'G': 15
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add notes section
    current_row += 2
    ws.merge_cells(f'A{current_row}:G{current_row}')
    ws[f'A{current_row}'] = "IMPORTANT NOTES"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].alignment = center_alignment
    current_row += 1
    
    notes = [
        "• Always warm up for 5-10 minutes before training",
        "• Cool down and stretch after each session",
        "• Stay hydrated throughout your workouts",
        "• Listen to your body - adjust intensity if needed",
        "• Track your progress weekly",
        "• Consult a healthcare provider before starting any new exercise program"
    ]
    
    for note in notes:
        ws[f'A{current_row}'] = note
        current_row += 1
    
    # Save the workbook
    filename = f"Training_Plan_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"Training plan saved as: {filename}")
    
    return filename

# Create a second worksheet for exercise library
def add_exercise_library(filename):
    wb = openpyxl.load_workbook(filename)
    
    # Create new worksheet for exercise library
    exercise_ws = wb.create_sheet("Exercise Library")
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Exercise data organized by category
    exercises = {
        "Strength - Upper Body": [
            {"exercise": "Push-ups", "sets": "3", "reps": "8-12", "rest": "60s", "notes": "Modify on knees if needed"},
            {"exercise": "Pull-ups/Lat Pulldowns", "sets": "3", "reps": "6-10", "rest": "90s", "notes": "Use assistance if needed"},
            {"exercise": "Overhead Press", "sets": "3", "reps": "8-10", "rest": "90s", "notes": "Dumbbells or barbell"},
            {"exercise": "Rows", "sets": "3", "reps": "10-12", "rest": "60s", "notes": "Cable, dumbbell, or barbell"},
            {"exercise": "Chest Press", "sets": "3", "reps": "8-12", "rest": "90s", "notes": "Bench press or dumbbells"}
        ],
        "Strength - Lower Body": [
            {"exercise": "Squats", "sets": "3", "reps": "12-15", "rest": "90s", "notes": "Bodyweight or weighted"},
            {"exercise": "Deadlifts", "sets": "3", "reps": "8-10", "rest": "120s", "notes": "Focus on form"},
            {"exercise": "Lunges", "sets": "3", "reps": "10 each leg", "rest": "60s", "notes": "Forward or reverse"},
            {"exercise": "Leg Press", "sets": "3", "reps": "12-15", "rest": "90s", "notes": "Machine exercise"},
            {"exercise": "Calf Raises", "sets": "3", "reps": "15-20", "rest": "45s", "notes": "Single or double leg"}
        ],
        "Core & Stability": [
            {"exercise": "Plank", "sets": "3", "reps": "30-60s", "rest": "60s", "notes": "Hold position"},
            {"exercise": "Dead Bug", "sets": "3", "reps": "10 each side", "rest": "45s", "notes": "Slow and controlled"},
            {"exercise": "Bird Dog", "sets": "3", "reps": "10 each side", "rest": "45s", "notes": "Hold 2-3 seconds"},
            {"exercise": "Side Plank", "sets": "2", "reps": "20-45s each", "rest": "60s", "notes": "Each side"},
            {"exercise": "Russian Twists", "sets": "3", "reps": "20", "rest": "45s", "notes": "With or without weight"}
        ],
        "Cardio Options": [
            {"exercise": "Running/Jogging", "sets": "1", "reps": "20-40 min", "rest": "-", "notes": "Adjust pace to intensity"},
            {"exercise": "Cycling", "sets": "1", "reps": "25-45 min", "rest": "-", "notes": "Indoor or outdoor"},
            {"exercise": "Swimming", "sets": "1", "reps": "20-35 min", "rest": "-", "notes": "Full body cardio"},
            {"exercise": "HIIT Circuit", "sets": "4-6", "reps": "30s work/30s rest", "rest": "2 min", "notes": "High intensity"},
            {"exercise": "Walking", "sets": "1", "reps": "30-60 min", "rest": "-", "notes": "Low impact option"}
        ]
    }
    
    current_row = 1
    
    # Title
    exercise_ws.merge_cells('A1:F1')
    exercise_ws['A1'] = "EXERCISE LIBRARY"
    exercise_ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    exercise_ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    exercise_ws['A1'].alignment = center_alignment
    current_row = 3
    
    for category, exercise_list in exercises.items():
        # Category header
        exercise_ws.merge_cells(f'A{current_row}:F{current_row}')
        exercise_ws[f'A{current_row}'] = category.upper()
        exercise_ws[f'A{current_row}'].font = header_font
        exercise_ws[f'A{current_row}'].fill = header_fill
        exercise_ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Column headers
        headers = ["Exercise", "Sets", "Reps/Duration", "Rest", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = exercise_ws.cell(row=current_row, column=col, value=header)
            cell.font = subheader_font
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1
        
        # Exercise data
        for exercise in exercise_list:
            row_data = [
                exercise['exercise'],
                exercise['sets'],
                exercise['reps'],
                exercise['rest'],
                exercise['notes']
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = exercise_ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                if col == 1:  # Exercise name
                    cell.font = Font(bold=True)
                elif col in [2, 3, 4]:  # Sets, Reps, Rest
                    cell.alignment = center_alignment
            current_row += 1
        
        current_row += 1  # Space between categories
    
    # Adjust column widths
    column_widths = {
        'A': 25,  # Exercise
        'B': 8,   # Sets
        'C': 15,  # Reps/Duration
        'D': 8,   # Rest
        'E': 35   # Notes
    }
    
    for col, width in column_widths.items():
        exercise_ws.column_dimensions[col].width = width
    
    wb.save(filename)
    print(f"Exercise library added to: {filename}")

# Run the functions
if __name__ == "__main__":
    # Create the main training plan
    filename = create_training_plan()
    
    # Add the exercise library
    add_exercise_library(filename)
    
    print("Complete training plan with exercise library created successfully!")
    print(f"File saved as: {filename}")